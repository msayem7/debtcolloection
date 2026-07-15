# from django.db import IntegrityError
import logging

# --------------------Organized Imports--------------------
# Standard Library Imports
import io
import json
from datetime import datetime, timedelta, date
from decimal import Decimal

# Django Imports
from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import connection, transaction, IntegrityError
from django.db.models import (
    F, Sum,Value, DecimalField,IntegerField, ExpressionWrapper, DurationField, DateField,
    Subquery, OuterRef, Q, Case, When, Func
)
from django.db.models.functions import Coalesce, Cast, Concat

from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.cache import never_cache
from django.utils.decorators import method_decorator
from django.utils import timezone
from django_filters import rest_framework as filters
# from django_filters import FilterSet, CharFilter, DateFilter, DecimalFilter
from django_filters.rest_framework import DjangoFilterBackend , FilterSet, CharFilter, DateFilter, NumberFilter


# Django REST Framework Imports
from rest_framework import viewsets, status, filters
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.viewsets import ViewSet
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.pagination import PageNumberPagination # Import pagination class

# Third-Party Imports
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from openpyxl import Workbook

# Local Application Imports
from .models import (
    Branch, Customer, CreditInvoice #, CustomerPayment, ChequeStore,
    #CustomerClaim, InvoiceChequeMap, InvoiceClaimMap, MasterClaim
)
from .models import PaymentInstrument, Payment, PaymentDetails, PaymentInstrumentType, Claim

from cheques import serializers
from .serializers import ( # You'll need to create these serializers
    ClaimListSerializer, ClaimUpdateSerializer
    #CustomerPaymentSerializer,  #ChequeStoreSerializer, CustomerClaimSerializer,
    # InvoiceChequeMapSerializer, MasterClaimSerializer
)
from .serializers import PaymentInstrumentSerializer, PaymentSerializer, PaymentDetailsSerializer


logger = logging.getLogger(__name__)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_detail(request):
    serializer = serializers.UserSerializer(request.user)
    return Response(serializer.data)

class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = serializers.CustomTokenObtainPairSerializer


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class BranchViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.BranchSerializer
    queryset = Branch.objects.all()
    permission_classes = [IsAuthenticated]
    lookup_field = 'alias_id'

    def update(self, request, *args, **kwargs):
        with transaction.atomic():
            client_version = int(request.data.get('version'))
            instance = self.get_object()

            # Concurrency check
            if instance.version != client_version:
                return Response(
                    {'version': 'This branch has been modified by another user. Please refresh. current V, client_version v: ' + str(instance.version) + ' ' + str(client_version)},
                    status=status.HTTP_409_CONFLICT
                )

            # Increment version
            new_version = instance.version + 1

            # Partial update handling
            partial = kwargs.pop('partial', False)
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)

            # Save with updated information
            serializer.save(updated_by=request.user, version=new_version)

            return Response(serializer.data)

    def perform_create(self, serializer):
        serializer.save(updated_by=self.request.user)

class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = serializers.CustomerSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'alias_id'
    filterset_fields = ['is_parent', 'parent']

    def get_queryset(self):
        queryset = super().get_queryset()
        branch_id = self.request.query_params.get('branch')        
         
        #if not self.request.user.is_staff:  # Example: admins see all
        if self.request.query_params.get('is_active'):
            is_active = self.request.query_params.get('is_active', 'true').lower() == 'true'
            # print('is_active',is_active, 'self.request.query_params.get', self.request.query_params.get('is_active', 'true').lower())
            queryset = queryset.filter(is_active=is_active)
        
        # Filter by branch alias_id
        if branch_id:
            queryset = queryset.filter(branch__alias_id=branch_id)
            
        # Filter parent customers
        if self.request.query_params.get('is_parent'):
            is_parent = self.request.query_params.get('is_parent', 'true').lower() == 'true'
            queryset = queryset.filter(is_parent=is_parent)

        queryset = queryset.annotate(
            sort_order=Case(
                When(parent__name__isnull=True, then=F('name')),
                default=Concat('parent__name', 'name')
            )
        ).order_by('sort_order')
        
        # print('queryset :', print(str(queryset.query)))
        return queryset
    

    def update(self, request, *args, **kwargs):
        try:
            is_active = request.data.get('is_active', None)
            if not is_active and (HasCustomerActivity.has_Activity(self, request)):
                return  Response({'error': 'Customer has active invoices or cheques. Inactivation is not possible'}, status=status.HTTP_409_CONFLICT)
            return super().update(request, *args, **kwargs)
        except Exception as e:
            print("Error:", e)
            return  Response({"error": f"Customer has active invoices. Inactivation is not possible. {e}"}, status=status.HTTP_409_CONFLICT)

class HasCustomerActivity(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = serializers.CustomerSerializer

    def has_Activity(self, request, *args, **kwargs):
        try:
            customer = get_object_or_404(Customer, alias_id=request.parser_context['kwargs']['alias_id'])
            if customer.is_parent:
                return (
                    CreditInvoice.objects.filter(Q(customer__parent=customer, payment__isnull=True)).exists()
                )
            else:
                return (
                    CreditInvoice.objects.filter(customer=customer, payment__isnull=True).exists() 
                )
            # return has_activity
        except Customer.DoesNotExist:
            return False
        
class CreditInvoiceViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.CreditInvoiceSerializer
    queryset = CreditInvoice.objects.select_related(
        'branch', 'customer', 'payment'
    ).all()
    lookup_field = 'alias_id'
    
    class payment:
        PAID = 'paid'
        UNPAID = 'unpaid'
        All = 'all'

    
    def get_queryset(self):
        params = self.request.query_params
        branch = params.get('branch')
        customer = params.get('customer')
        date_from = params.get('transaction_date_after')
        date_to = params.get('transaction_date_before')
        payment_status = params.get('payment', 'all')
        report_date = params.get('report_date')

        queryset = CreditInvoice.objects.select_related(
            'branch', 'customer', 'payment'
        ).all()
        # Apply filters
        if branch:
            queryset = queryset.filter(branch__alias_id=branch)
        if date_from:
            queryset = queryset.filter(transaction_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(transaction_date__lte=date_to)

        if customer:
            cust = Customer.objects.only('id', 'is_parent').filter(alias_id=customer).first()
            if cust:
                if cust.is_parent:
                    # Use subquery instead of materializing child IDs in Python
                    queryset = queryset.filter(
                        customer__in=Customer.objects.filter(parent=cust).values('pk')
                    )
                else:
                    queryset = queryset.filter(customer=cust)

        # Handle payment status filter
        if payment_status.lower() == 'unpaid':
            queryset = queryset.filter(payment__isnull=True)
            # Handle matured dues if report_date is provided
            if report_date:
                try:
                    report_date_obj = datetime.strptime(report_date, '%Y-%m-%d').date()
                    # Filter on indexed transaction_date first to reduce rows,
                    # then apply the grace date expression
                    queryset = queryset.filter(
                        transaction_date__lte=report_date_obj
                    ).extra(
                        where=["transaction_date + payment_grace_days <= %s"],
                        params=[report_date_obj]
                    )
                except ValueError:
                    pass  # Ignore invalid date format
        elif payment_status.lower() == 'paid' or payment_status.lower() == 'all':
                pass
        elif payment_status:
                queryset = queryset.filter(payment__alias_id=payment_status)
        
        return queryset.order_by('transaction_date')

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()

        if instance.payment and 'customer' in request.data:
            return Response(
                {'error': 'Cannot change customer after invoice has been paid'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if int(request.data.get('version')) != instance.version:
            return Response({'error': 'Version conflict'}, status=status.HTTP_409_CONFLICT)

        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(
            instance, 
            data=request.data,  
            partial=partial
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(updated_by=request.user, version=instance.version + 1)

        return Response(serializer.data)

    @method_decorator(never_cache)  # Disable caching
    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if latest := CreditInvoice.objects.only('updated_at').order_by('-updated_at').first():
            response.headers['Last-Modified'] = latest.updated_at.strftime('%a, %d %b %Y %H:%M:%S GMT')
        return response

# payment implemente here 

class PaymentInstrumentTypeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PaymentInstrumentType.objects.all()
    serializer_class = serializers.PaymentInstrumentTypeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        branch_id = self.request.query_params.get('branch')

        if branch_id:
            # Use alias_id directly in the filter
            queryset = queryset.filter(branch__alias_id=branch_id)

        return queryset.order_by('serial_no')
    
    
class PaymentInstrumentsViewSet(viewsets.ModelViewSet):
    queryset = PaymentInstrument.objects.all()
    serializer_class = PaymentInstrumentSerializer
    # Remove filterset_fields since we'll handle filtering manually
    # filterset_fields= ['branch', 'is_active']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        branch_id = self.request.query_params.get('branch')
        instrument_type_serial_no = self.request.query_params.get('instrument_type_serial_no')
        is_active = self.request.query_params.get('is_active', 'true').lower() == 'true'
        
        
        queryset = queryset.filter(is_active=is_active)

        if branch_id:
            # Use alias_id directly in the filter
            queryset = queryset.filter(branch__alias_id=branch_id)

        if instrument_type_serial_no:
            queryset = queryset.filter(instrument_type__serial_no=instrument_type_serial_no)

        return queryset.order_by('serial_no')
    

class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer #PaymentViewSerializer
    lookup_field = 'alias_id'
    
    def get_serializer_class(self):
        # if self.action == 'create':
        #     return PaymentSerializer
        return PaymentSerializer #PaymentViewSerializer
     
    def get_queryset(self):
        queryset = super().get_queryset()

        branch_id = self.request.query_params.get('branch')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        customer_id = self.request.query_params.get('customer')
        # is_fully_allocated = self.request.query_params.get('is_fully_allocated')
        
    
        if branch_id:
            queryset = queryset.filter(branch__alias_id=branch_id)
            
        if date_from:
            queryset = queryset.filter(received_date__gte=date_from)
            
        if date_to:
            queryset = queryset.filter(received_date__lte=date_to)
            
        if customer_id:
            queryset = queryset.filter(customer__alias_id=customer_id)
       
        # print (queryset.query)
        
        return queryset.order_by('-received_date')
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        # Get data from the request
        validated_data = request.data
        
        # Extract nested data
        payment_details_data = validated_data.pop('payment_details', [])
        invoices_data = validated_data.pop('invoices', [])
        cash_equivalent_amount = validated_data.pop('cash_equivalent_amount', 0.0)
        total_amount = validated_data.pop('total_amount', 0.0)
        shortage_amount = validated_data.pop('shortage_amount', 0.0)

       
        branch_alias_id = validated_data.get('branch')
        try:
            branch = Branch.objects.get(alias_id=branch_alias_id)
        except Branch.DoesNotExist:
            return Response({"error": f"Branch with alias_id {branch_alias_id} does not exist."}, status=status.HTTP_400_BAD_REQUEST)
        validated_data['branch'] = branch

        customer_alias_id = validated_data.get('customer')
        try:
            customer = Customer.objects.get(alias_id=customer_alias_id)
        except Customer.DoesNotExist:
            return Response({"error": f"Customer with alias_id {customer_alias_id} does not exist."}, status=status.HTTP_400_BAD_REQUEST)
        validated_data['customer'] = customer
         # Create the Payment instance
        payment = Payment.objects.create(**validated_data)

        # Handle PaymentDetails
        errors = {}
        for index, detail_data in enumerate(payment_details_data):
            payment_instrument = detail_data['payment_instrument']
            
            if 'alias_id' in detail_data and not detail_data['alias_id']:
                del detail_data['alias_id']

            try:
                instrument = PaymentInstrument.objects.get(id=payment_instrument)
            except Customer.DoesNotExist:
                return Response({"error": f"Instruement with id {payment_instrument} does not exist."}, status=status.HTTP_400_BAD_REQUEST)
            payment_details_data[index]['payment_instrument'] = instrument
            
            # Handle auto-number generation
            if instrument.instrument_type.auto_number:
                locked_type = PaymentInstrumentType.objects.select_for_update().get(pk=instrument.instrument_type.id)
                locked_type.last_number += 1
                detail_data['id_number'] = f"{locked_type.prefix}{locked_type.last_number:04d}"
                locked_type.save()
            else:
                # Check if the ID number is unique within the same branch
                if PaymentDetails.objects.filter(branch=payment.branch, id_number=detail_data.get('id_number')).exists():
                    errors[f'payment_details.{index}.id_number'] = ["This ID number already exists in this branch."]
        
        if errors:
            return Response({"errors": errors}, status=status.HTTP_400_BAD_REQUEST)

        # Create PaymentDetails and claim objects
        for detail_data in payment_details_data:
            
            payment_details= PaymentDetails.objects.create(payment=payment, branch=payment.branch, **detail_data)
            if instrument.instrument_type.serial_no == 3:
                Claim.objects.create(branch=payment.branch, payment_details = payment_details
                )

        # Update CreditInvoices
        for invoice_data in invoices_data:
            invoice_alias_id = invoice_data.get('alias_id')
            if not invoice_alias_id:
                return Response({"error": "Missing 'alias_id' for one or more invoices"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                invoice = CreditInvoice.objects.get(alias_id=invoice_alias_id)
                invoice.payment = payment # Mark invoice as paid
                invoice.status = True  
                invoice.save()
            except CreditInvoice.DoesNotExist:
                return Response({"error": f"Invoice with alias_id {invoice_alias_id} does not exist."}, status=status.HTTP_400_BAD_REQUEST)

        # Update the Payment amounts
        payment.total_amount = total_amount
        payment.cash_equivalent_amount = cash_equivalent_amount
        payment.shortage_amount = shortage_amount
        payment.save()

        # Return the created payment object with the serializer
        # return Response(PaymentViewSerializer(payment).data, status=status.HTTP_201_CREATED)
        return Response(PaymentSerializer(payment).data, status=status.HTTP_201_CREATED)
    

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        payment = self.get_object()
        
        # Version check
        client_version = request.data.get('version')
        if client_version and int(client_version) != payment.version:
            return Response(
                {"error": "This payment has been modified by another user. Please refresh."},
                status=status.HTTP_409_CONFLICT
            )

        # Extract data
        validated_data = request.data.copy()
        payment_details_data = validated_data.pop('payment_details', [])
        invoices_data = validated_data.pop('invoices', [])
        print('payment_details_data:', payment_details_data)

        # Update payment fields
        for field, value in validated_data.items():
            if field in ['branch', 'customer']:
                try:
                    if field == 'branch':
                        obj = Branch.objects.get(alias_id=value)
                    else:
                        obj = Customer.objects.get(alias_id=value)
                    setattr(payment, field, obj)
                except (Branch.DoesNotExist, Customer.DoesNotExist) as e:
                    return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            elif hasattr(payment, field):
                setattr(payment, field, value)

        # Handle payment details updates
        #existing_detail_alias_ids = [d.alias_id for d in payment.paymentdetails_set.all()]
        # request_detail_alias_ids = [d.get('alias_id') for d in payment_details_data if d.get('alias_id')]
        updated_detail_ids = []

        # if not set(existing_detail_id_numbers).issubset(set(request_detail_id_numbers)):
        #     return Response({"error": "Numbers of existing information can't be deleted removed"}, status=status.HTTP_400_BAD_REQUEST)
        
        # print('request_detail_id_numbers:', request_detail_id_numbers)

        for detail_data in payment_details_data:
            detail_alias_id = detail_data.get('alias_id')
            if detail_alias_id: # and detail_alias_id in existing_detail_alias_ids:
                # Update existing detail
                try:
                    detail = PaymentDetails.objects.get(alias_id=detail_alias_id)
                    # instrument = PaymentInstrument.objects.get(id=detail_data['payment_instrument'])
                    if not detail.payment_instrument.id == detail_data['payment_instrument'] or not detail.id_number == detail_data['id_number']:
                        return Response({"error": "Numbers of existing instrument or Id Number can't be deleted or changed"}, status=status.HTTP_400_BAD_REQUEST)
                        # No changes to instrument or ID number, just update other fields
                    # For existing details, don't change ID number
                    for field, value in detail_data.items():
                        # if field == 'payment_instrument':
                        #     setattr(detail, field, instrument)
                        # elif 
                        if field != 'id_number' and field != 'payment_instrument' and field != 'alias_id' and hasattr(detail, field):
                            setattr(detail, field, value)
                    detail.save()
                    updated_detail_ids.append(detail.alias_id)
                except (PaymentDetails.DoesNotExist, PaymentInstrument.DoesNotExist):
                    continue
            else:
                # Create new detail
                try:
                    instrument = PaymentInstrument.objects.get(id=detail_data['payment_instrument'])
                    id_number = None
                    
                    # Handle auto-numbering
                    if instrument.instrument_type.auto_number:
                        # Lock the type row to prevent concurrent updates
                        # with transaction.atomic():
                        locked_type = PaymentInstrumentType.objects.select_for_update().get(
                            pk=instrument.instrument_type.id
                        )
                        locked_type.last_number += 1
                        id_number = f"{locked_type.prefix}{locked_type.last_number:04d}"
                        locked_type.save()
                    else:
                        id_number = detail_data.get('id_number', '')
                        # Manual ID - check uniqueness
                        if id_number and PaymentDetails.objects.filter(
                            branch=payment.branch, 
                            id_number=id_number
                        ).exists():
                            return Response(
                                {"error": f"ID number {id_number} already exists in this branch"},
                                status=status.HTTP_400_BAD_REQUEST
                            )
                    
                    # Create detail
                    detail = PaymentDetails.objects.create(
                        payment=payment,
                        branch=payment.branch,
                        payment_instrument=instrument,
                        id_number=id_number,
                        amount=detail_data.get('amount', 0),
                        detail=detail_data.get('detail', '')
                    )
                    updated_detail_ids.append(detail.alias_id)
                    
                    # Create claim if needed
                    if instrument.instrument_type.serial_no == 3:
                        Claim.objects.create(
                            branch=payment.branch,
                            payment_details=detail
                        )
                            
                except PaymentInstrument.DoesNotExist:
                    continue

        # Delete removed details
        if (PaymentDetails.objects.filter(payment=payment).exclude(alias_id__in=updated_detail_ids).exists()):
            return Response(
                {"error": "Cannot remove existing payment details."},
                status=status.HTTP_400_BAD_REQUEST
            )
        # PaymentDetails.objects.filter(payment=payment).exclude(id__in=updated_detail_ids).delete()

        # Handle invoice updates
        existing_invoice_ids = [i.alias_id for i in payment.invoice_set.all()]
        updated_invoice_ids = []
        
        for invoice_data in invoices_data:
            invoice_alias_id = invoice_data.get('alias_id')
            if invoice_alias_id:
                try:
                    invoice = CreditInvoice.objects.get(alias_id=invoice_alias_id)
                    invoice.payment = payment
                    invoice.status = True
                    invoice.save()
                    updated_invoice_ids.append(invoice.alias_id)
                except CreditInvoice.DoesNotExist:
                    continue

        # Unlink removed invoices
        CreditInvoice.objects.filter(
            payment=payment
        ).exclude(
            alias_id__in=updated_invoice_ids
        ).update(
            payment=None,
            status=False
        )

        # Update payment amounts and version
        payment.total_amount = validated_data.get('total_amount', 0)
        payment.cash_equivalent_amount = validated_data.get('cash_equivalent_amount', 0)
        payment.shortage_amount = validated_data.get('shortage_amount', 0)
        payment.version = F('version') + 1
        payment.save()
        payment.refresh_from_db()

        return Response(PaymentSerializer(payment).data)
    
    
        
# class ClaimFilter(FilterSet):
#     customer = CharFilter(field_name='payment_details__payment__customer__alias_id', lookup_expr='icontains')
#     instrument = CharFilter(field_name='payment_details__payment_instrument__serial_no', lookup_expr='exact')
#     claim_date = DateFilter(field_name='payment_details__payment__received_date', lookup_expr='gte')

#     # Range filters for claim_amount, refund_amount, and remaining_amount
#     claim_amount_min = NumberFilter(field_name='payment_details__amount', lookup_expr='gte')  # claim_amount >= min
#     claim_amount_max = NumberFilter(field_name='payment_details__amount', lookup_expr='lte')  # claim_amount <= max
    
#     refund_amount_min = NumberFilter(field_name='refund_amount', lookup_expr='gte')  # refund_amount >= min
#     refund_amount_max = NumberFilter(field_name='refund_amount', lookup_expr='lte')  # refund_amount <= max
    
#     remaining_amount_min = NumberFilter(field_name='remaining_amount', lookup_expr='gte')  # remaining_amount >= min
#     remaining_amount_max = NumberFilter(field_name='remaining_amount', lookup_expr='lte')  # remaining_amount <= max
    

#     class Meta:
#         model = Claim
#         fields = ['customer', 'instrument', 'claim_date', 
#                   'claim_amount_min', 'claim_amount_max', 
#                   'refund_amount_min', 'refund_amount_max', 
#                   'remaining_amount_min', 'remaining_amount_max']

class ClaimFilter(FilterSet):
    customer = CharFilter(field_name='payment_details__payment__customer__alias_id', lookup_expr='icontains')
    instrument = CharFilter(field_name='payment_details__payment_instrument__id', lookup_expr='exact')
    claim_date = DateFilter(field_name='payment_details__payment__received_date', lookup_expr='gte')

    # Range filters for claim_amount, refund_amount, and remaining_amount
    claim_amount_min = NumberFilter(field_name='payment_details__amount', lookup_expr='gte')
    claim_amount_max = NumberFilter(field_name='payment_details__amount', lookup_expr='lte')
    
    refund_amount_min = NumberFilter(field_name='refund_amount', lookup_expr='gte')
    refund_amount_max = NumberFilter(field_name='refund_amount', lookup_expr='lte')
    
    remaining_amount_min = NumberFilter(field_name='remaining_amount', lookup_expr='gte')
    remaining_amount_max = NumberFilter(field_name='remaining_amount', lookup_expr='lte')

    class Meta:
        model = Claim
        fields = [
            'customer', 
            'instrument', 
            'claim_date',
            'claim_amount_min', 
            'claim_amount_max',
            'refund_amount_min', 
            'refund_amount_max',
            'remaining_amount_min', 
            'remaining_amount_max'
        ]

class ClaimViewSet(viewsets.ModelViewSet):
    queryset = Claim.objects.select_related(
        'payment_details__payment__customer',
        'payment_details__payment_instrument',
        'payment_details__payment_instrument__instrument_type',
    ).filter(
        payment_details__payment_instrument__instrument_type__serial_no=3
    ).annotate(
        remaining_amount=ExpressionWrapper(
            F('payment_details__amount') - Coalesce(F('refund_amount'), 0),
            output_field=DecimalField()
        )
    )  # Assuming serial_no 3 is for claims
    # queryset =queryset.filter(instrument_type__serial_no=3)  # Assuming serial_no 3 is for claims
    serializer_class = ClaimListSerializer
    lookup_field = 'alias_id'

    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = ClaimFilter
    # filterset_fields = {
    #     'payment_details__payment__customer__alias_id': ['icontains'],
    #     'payment_details__payment_instrument__instrument_name': ['icontains'],
    #     'payment_details__payment__received_date': ['lte'], #['exact', 'gte', 'lte'],
    # }
    search_fields = [
        'payment_details__payment__customer__alias_id',
        'payment_details__payment_instrument__instrument_name',
    ]
    
    # Add this line to enable pagination for this ViewSet
    # pagination_class = StandardResultsSetPagination

    # def get_queryset(self):
    #     queryset = super().get_queryset()

    #     branch_id = self.request.query_params.get('branch')
    #     date_from = self.request.query_params.get('date_from')
    #     date_to = self.request.query_params.get('date_to')
    #     customer_id = self.request.query_params.get('customer')
    #     # is_fully_allocated = self.request.query_params.get('is_fully_allocated')
        
    
    #     if branch_id:
    #         queryset = queryset.filter(branch__alias_id=branch_id)
            
    #     if date_from:
    #         queryset = queryset.filter(received_date__gte=date_from)
            
    #     if date_to:
    #         queryset = queryset.filter(received_date__lte=date_to)
            
    #     if customer_id:
    #         queryset = queryset.filter(customer__alias_id=customer_id)
       
    #     # print (queryset.query)
        
    #     return queryset.order_by('-received_date')

    def get_serializer_class(self):
        if self.action in ['update', 'partial_update']:
            return ClaimUpdateSerializer
        return ClaimListSerializer

    @action(detail=True, methods=['patch'], permission_classes=[IsAuthenticated])
    def update_claim(self, request, *args, **kwargs):
        claim = self.get_object()
        serializer = self.get_serializer(claim, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save(updated_by=request.user, version=claim.version + 1)
        return Response(serializer.data)
    
# ----------------- end of payment implementation---------------------


# --------Latest:01  parent customer due
class ParentCustomerDueReport(APIView):
    def get(self, request):
        report_date_str = request.query_params.get('date')
        branch_alias_id = request.query_params.get('branch')

        if branch_alias_id is None:
            return Response(
                {"error": "Banch Id is mandatory"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            report_date = timezone.datetime.strptime(report_date_str, '%Y-%m-%d').date() if report_date_str else timezone.now().date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )

        customer_qs = Customer.objects.all()
        invoice_qs = CreditInvoice.objects.all()

        # Apply branch filter if provided
        if branch_alias_id:
            customer_qs = customer_qs.filter(branch__alias_id=branch_alias_id)
            invoice_qs = invoice_qs.filter(branch__alias_id=branch_alias_id)
        else:
            return Response(
                {"error": "Branch is mandatory"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Query 1: Get all parent-child relationships with alias_id
        customer_hierarchy = customer_qs.filter(
            ( Q(is_parent=True) | Q(parent__isnull=False))
        ).values('alias_id', 'name', 'is_parent', 'parent__alias_id', 'parent__name')

        # Query 2: Get all due invoice amounts grouped by customer
        due_amounts = invoice_qs.filter(
            customer__parent__isnull=False,  # Only child customers
            transaction_date__lte=report_date
        ).filter(
            Q(payment__isnull=True) |
            Q(payment__received_date__gt=report_date)
        ).annotate(
            is_matured=Case(
                When(transaction_date__lte=report_date-F('payment_grace_days'),
                     then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            )
        ).values('customer__alias_id').annotate(
            matured_due=Sum('sales_amount', filter=Q(is_matured=1)) - Sum('sales_return', filter=Q(is_matured=1)),
            immature_due=Sum('sales_amount', filter=Q(is_matured=0))- Sum('sales_return', filter=Q(is_matured=0))
        )

        # Check if export is requested
        output = request.query_params.get('export', 'html').lower()
        if output == 'excel':
            return self._export_excel(request, report_date, branch_alias_id, customer_hierarchy, due_amounts)

        # Process in Python for JSON response
        parents = {c['alias_id']: c for c in customer_hierarchy if c['is_parent']}
        children = {c['alias_id']: c for c in customer_hierarchy if not c['is_parent']}

        report_data = []
        grand_total_matured = 0
        grand_total_immature = 0
        grand_total_due = 0

        for parent in parents.values():
            parent_entry = {
                'alias_id': parent['alias_id'],
                'name': parent['name'],
                'matured_due': 0,
                'immature_due': 0,
                'total_due': 0,
                'children': []
            }

            for child in children.values():
                if child['parent__alias_id'] == parent['alias_id']:
                    amounts = next(
                        (a for a in due_amounts
                         if a['customer__alias_id'] == child['alias_id']),
                        {}
                    )
                    child_entry = {
                        'alias_id': child['alias_id'],
                        'name': child['name'],
                        'matured_due': amounts.get('matured_due', Decimal(0)) or Decimal(0),
                        'immature_due': amounts.get('immature_due', Decimal(0)) or Decimal(0)
                    }
                    child_entry['total_due'] = child_entry['matured_due'] + child_entry['immature_due']
                    parent_entry['children'].append(child_entry)
                    parent_entry['matured_due'] += child_entry['matured_due']
                    parent_entry['immature_due'] += child_entry['immature_due']
                    parent_entry['total_due'] += child_entry['total_due']

                    grand_total_matured += child_entry['matured_due']
                    grand_total_immature += child_entry['immature_due']
                    grand_total_due += child_entry['total_due']

            report_data.append(parent_entry)

        return Response({
            'report_date': report_date.strftime('%Y-%m-%d'),
            'data': report_data,
            'grand_totals': {
                'matured_due': grand_total_matured,
                'immature_due': grand_total_immature,
                'total_due': grand_total_due
            }
        })

    def _export_excel(self, request, report_date, branch_alias_id, customer_hierarchy, due_amounts):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # Process data
        parents = {c['alias_id']: c for c in customer_hierarchy if c['is_parent']}
        children = {c['alias_id']: c for c in customer_hierarchy if not c['is_parent']}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Customer Due Report'

        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        title_font = Font(name='Calibri', bold=True, size=14, color='2F5496')
        label_font = Font(name='Calibri', bold=True, size=10)
        value_font = Font(name='Calibri', size=10)

        thin_border = Border(
            left=Side(style='thin', color='B0B0B0'),
            right=Side(style='thin', color='B0B0B0'),
            top=Side(style='thin', color='B0B0B0'),
            bottom=Side(style='thin', color='B0B0B0'),
        )
        number_alignment = Alignment(horizontal='right', vertical='center')
        text_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = 'Parent Customer Due Report'
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        # Filter info
        try:
            branch_name = Branch.objects.values_list('name', flat=True).get(
                alias_id=branch_alias_id
            )
        except Branch.DoesNotExist:
            branch_name = branch_alias_id

        filter_info = [
            ('Report Date', report_date.strftime('%Y-%m-%d')),
            ('Branch', branch_name),
        ]
        for i, (label, value) in enumerate(filter_info):
            row_num = 2 + i
            ws.cell(row=row_num, column=1, value=label).font = label_font
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=4)
            val_cell = ws.cell(row=row_num, column=2, value=value)
            val_cell.font = value_font
            val_cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.append([])

        # Column headers
        cols = ['Customer', 'Matured', 'Immature', 'Total']
        header_row = ws.max_row + 1
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 28

        # Data rows
        data_start_row = header_row + 1
        row_num = data_start_row

        for parent in parents.values():
            # Parent row
            parent_matured = 0
            parent_immature = 0
            parent_total = 0
            child_rows = []

            for child in children.values():
                if child['parent__alias_id'] == parent['alias_id']:
                    amounts = next(
                        (a for a in due_amounts
                         if a['customer__alias_id'] == child['alias_id']),
                        {}
                    )
                    matured = float(amounts.get('matured_due', 0) or 0)
                    immature = float(amounts.get('immature_due', 0) or 0)
                    total = matured + immature

                    if total != 0:
                        child_rows.append((child['name'], matured, immature, total))
                        parent_matured += matured
                        parent_immature += immature
                        parent_total += total

            if parent_total != 0 or len(child_rows) > 0:
                # Write parent row
                cells_data = [
                    (1, parent['name'], text_alignment, True),
                    (2, parent_matured, number_alignment, True),
                    (3, parent_immature, number_alignment, True),
                    (4, parent_total, number_alignment, True),
                ]
                for col_idx, val, align, is_bold in cells_data:
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.font = Font(name='Calibri', bold=True, size=10)
                    cell.alignment = align
                    cell.border = thin_border
                row_num += 1

                # Write child rows
                for child_name, matured, immature, total in child_rows:
                    child_cells = [
                        (1, f'  {child_name}', text_alignment, False),
                        (2, matured, number_alignment, False),
                        (3, immature, number_alignment, False),
                        (4, total, number_alignment, False),
                    ]
                    for col_idx, val, align, _ in child_cells:
                        cell = ws.cell(row=row_num, column=col_idx, value=val)
                        cell.font = value_font
                        cell.alignment = align
                        cell.border = thin_border
                    row_num += 1

        # Grand total row
        grand_matured = 0
        grand_immature = 0
        grand_total = 0
        for parent in parents.values():
            for child in children.values():
                if child['parent__alias_id'] == parent['alias_id']:
                    amounts = next(
                        (a for a in due_amounts
                         if a['customer__alias_id'] == child['alias_id']),
                        {}
                    )
                    grand_matured += float(amounts.get('matured_due', 0) or 0)
                    grand_immature += float(amounts.get('immature_due', 0) or 0)
                    grand_total += float(amounts.get('matured_due', 0) or 0) + float(amounts.get('immature_due', 0) or 0)

        grand_total_cells = [
            (1, 'Grand Total', text_alignment, True),
            (2, grand_matured, number_alignment, True),
            (3, grand_immature, number_alignment, True),
            (4, grand_total, number_alignment, True),
        ]
        for col_idx, val, align, _ in grand_total_cells:
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            cell.alignment = align
            cell.border = thin_border

        # Column widths
        col_widths = {1: 40, 2: 16, 3: 16, 4: 16}
        for col_idx in range(1, 5):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 14)

        ws.freeze_panes = f'A{header_row + 1}'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        wb.save(response)
        response['Content-Disposition'] = 'attachment; filename="parent_customer_due_report.xlsx"'
        return response
# ==================== Credit Invoice Report Module ====================

class CreditInvoiceReportView(APIView):
    '''
    Credit Invoice Report with filtering, pagination, PDF, Excel, CSV exports.
    Endpoint: GET /v1/chq/reports/invoice/
    '''
    permission_classes = [IsAuthenticated]

    # Query 1: Base Query
    QUERY1 = '''
        SELECT
            ROW_NUMBER() OVER (
                ORDER BY  p.received_date ASC NULLS LAST, ci.payment_id, ci.transaction_date 
            ) AS sl_no, 
            pc.name AS parent_organization,
            c.name AS customer_name,
            ci.transaction_date,
            ci.payment_grace_days AS grace_days,
            (ci.transaction_date + ci.payment_grace_days) Due_Date,
            p.received_date,
            ci.sales_amount,
            ci.sales_return,
            ci.sales_amount - ci.sales_return AS net_sales,
            p.cash_equivalent_amount AS cheque_cash_amount,
            p.claim_amount,
            p.shortage_amount,
            GREATEST(
                COALESCE(p.received_date, %(report_date)s::date) - (ci.transaction_date + ci.payment_grace_days),
                0
            ) AS days_overdue,
            p.alias_id payment_alias_id,
            ci.payment_id as payment_SL
        FROM credit_invoice ci
        LEFT JOIN payment p ON p.id = ci.payment_id
        LEFT JOIN customer c ON c.id = ci.customer_id
        LEFT JOIN customer pc ON pc.id = c.parent_id AND pc.is_parent = TRUE
        WHERE ci.branch_id = %(branch_id)s
        {status_filter}
        {parent_filter}
        {child_filter}
        {date_filter}
        {return_filter}
        ORDER BY p.received_date ASC NULLS LAST, ci.payment_id, ci.transaction_date
    '''
    # ORDER BY p.received_date DESC NULLS LAST, ci.payment_id, ci.transaction_date

    # Query 2: Detailed Query (with instrument numbers)
    QUERY2 = '''
        WITH payment_summary AS (
            SELECT
                pd.payment_id,
                STRING_AGG(pd.id_number, '; ') FILTER (WHERE pit.is_cash_equivalent) AS cheque_numbers,
                SUM(pd.amount) FILTER (WHERE pit.is_cash_equivalent) AS cheque_cash_amount,
                STRING_AGG(pd.id_number, '; ') FILTER (WHERE pit.is_cash_equivalent = FALSE) AS claim_numbers,
                SUM(pd.amount) FILTER (WHERE pit.is_cash_equivalent = FALSE) AS claim_amount
            FROM payment_details pd
            JOIN payment_instrument pi ON pi.id = pd.payment_instrument_id
            JOIN payment_instrument_type pit ON pit.id = pi.instrument_type_id
            GROUP BY pd.payment_id
        ),
        invoices_ranked AS (
            SELECT
                pc.name AS parent_organization,
                c.name AS customer_name,                
                ci.customer_id, ci.transaction_date, ci.payment_grace_days,
                ci.sales_amount, ci.sales_return,
                (ci.sales_amount - ci.sales_return) AS net_sales,
                p.received_date, ci.payment_id, p.shortage_amount,
                p.alias_id payment_alias_id,
                ROW_NUMBER() OVER (
                    PARTITION BY ci.payment_id
                    ORDER BY p.received_date ASC NULLS LAST, ci.payment_id, ci.transaction_date
                ) AS serial_num
            FROM credit_invoice ci
            LEFT JOIN payment p ON p.id = ci.payment_id
            LEFT JOIN customer c ON c.id = ci.customer_id
            LEFT JOIN customer pc ON pc.id = c.parent_id AND pc.is_parent = TRUE
            WHERE ci.branch_id = %(branch_id)s
            {status_filter}
            {parent_filter}
            {child_filter}
            {date_filter}
            {return_filter}
        )
        SELECT
            ROW_NUMBER() OVER (
                ORDER BY  ci.received_date ASC NULLS LAST, ci.payment_id, ci.serial_num, ci.transaction_date 
            ) AS sl_no, 
	        serial_num,
            ci.parent_organization,
            ci.customer_name,
            ci.transaction_date,
            (ci.transaction_date + ci.payment_grace_days) Due_Date,
            ci.received_date,
            ci.sales_amount,
            ci.sales_return,
            ci.net_sales,
            ps.cheque_numbers,
            ps.cheque_cash_amount,
            ps.claim_numbers,
            ps.claim_amount,
            case when ci.serial_num = 1 
                then  ci.shortage_amount 
                else 0 
                end as shortage_amount,
            GREATEST(
                COALESCE(ci.received_date, %(report_date)s::date) - (ci.transaction_date + ci.payment_grace_days),
                0
            ) AS days_overdue,
            ci.payment_grace_days as grace_days,
            ci.payment_alias_id,
            ci.serial_num as payment_SL
        FROM invoices_ranked ci
        LEFT JOIN payment_summary ps
            ON ps.payment_id = ci.payment_id AND ci.serial_num = 1
        ORDER BY sl_no, ci.serial_num, ci.transaction_date
    '''
    # ORDER BY ci.received_date DESC NULLS LAST, ci.payment_id, ci.serial_num, ci.transaction_date


    # Count Query (for pagination)
    COUNT_QUERY = '''
        SELECT COUNT(*) AS total
        FROM credit_invoice ci
        LEFT JOIN payment p ON p.id = ci.payment_id
        LEFT JOIN customer c ON c.id = ci.customer_id
        LEFT JOIN customer pc ON pc.id = c.parent_id AND pc.is_parent = TRUE
        WHERE ci.branch_id = %(branch_id)s
        {status_filter}
        {parent_filter}
        {child_filter}
        {date_filter}
        {return_filter}
    '''

    # Totals Query (for report footer)
    TOTALS_QUERY = '''
        SELECT
            COALESCE(SUM(ci.sales_amount), 0) AS total_sales_amount,
            COALESCE(SUM(ci.sales_return), 0) AS total_sales_return,
            COALESCE(SUM(ci.sales_amount - ci.sales_return), 0) AS total_net_sales
        FROM credit_invoice ci
        LEFT JOIN payment p ON p.id = ci.payment_id
        LEFT JOIN customer c ON c.id = ci.customer_id
        LEFT JOIN customer pc ON pc.id = c.parent_id AND pc.is_parent = TRUE
        WHERE ci.branch_id = %(branch_id)s
        {status_filter}
        {parent_filter}
        {child_filter}
        {date_filter}
        {return_filter}
    '''

    # Helper: Build filter snippets
    def _build_filters(self, params):
        snippets = {
            'status_filter': '',
            'parent_filter': '',
            'child_filter': '',
            'date_filter': '',
            'return_filter': '',
        }
        extra_params = {}

        status = params.get('status', 'all')
        if status == 'due':
            snippets['status_filter'] = ' AND ci.payment_id IS NULL'
        elif status == 'immature_due':
            snippets['status_filter'] = (
                " AND ci.payment_id IS NULL"
                " AND (ci.transaction_date + ci.payment_grace_days) >= %(report_date)s::date"
            )
            extra_params['report_date'] = params['report_date']
        elif status == 'matured_due':
            snippets['status_filter'] = (
                " AND ci.payment_id IS NULL"
                " AND (ci.transaction_date + ci.payment_grace_days) < %(report_date)s::date"
            )
            extra_params['report_date'] = params['report_date']
        elif status == 'paid':
            snippets['status_filter'] = ' AND ci.payment_id IS NOT NULL'

        # Customer hierarchy
        parent_customer = params.get('parent_customer')
        child_customer = params.get('child_customer')

        if child_customer:
            snippets['child_filter'] = " AND c.alias_id = %(child_customer)s"
            extra_params['child_customer'] = child_customer
        elif parent_customer:
            snippets['parent_filter'] = " AND pc.alias_id = %(parent_customer)s"
            extra_params['parent_customer'] = parent_customer

        # Date filter - Received date mode only valid for 'all' and 'paid'
        date_mode = params.get('date_mode', 'transaction_date')
        if status in ('due', 'immature_due', 'matured_due'):
            date_mode = 'transaction_date'

        date_from = params.get('date_from')
        date_to = params.get('date_to')

        if date_from and date_to:
            col = 'p.received_date' if date_mode == 'received_date' else 'ci.transaction_date'
            snippets['date_filter'] = f" AND {col} BETWEEN %(date_from)s::date AND %(date_to)s::date"
            extra_params['date_from'] = date_from
            extra_params['date_to'] = date_to
        elif date_from:
            col = 'p.received_date' if date_mode == 'received_date' else 'ci.transaction_date'
            snippets['date_filter'] = f" AND {col} >= %(date_from)s::date"
            extra_params['date_from'] = date_from
        elif date_to:
            col = 'p.received_date' if date_mode == 'received_date' else 'ci.transaction_date'
            snippets['date_filter'] = f" AND {col} <= %(date_to)s::date"
            extra_params['date_to'] = date_to

        # Return only filter
        if params.get('return_only') in (True, 'true', 'True', 1, '1'):
            snippets['return_filter'] = ' AND ci.sales_return > 0'

        return snippets, extra_params

    def _get_common_params(self, request):
        branch = request.query_params.get('branch')
        if not branch:
            raise ValidationError({'branch': 'Branch is required.'})

        branch_obj = get_object_or_404(Branch, alias_id=branch)

        today = timezone.now().date()
        report_date = request.query_params.get('report_date', today.isoformat())

        try:
            report_date_obj = datetime.strptime(report_date, '%Y-%m-%d').date()
        except ValueError:
            raise ValidationError({'report_date': 'Invalid date format. Use YYYY-MM-DD.'})

        date_to = request.query_params.get('date_to')
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                if report_date_obj < date_to_obj:
                    raise ValidationError(
                        {'report_date': 'Report date must be >= To date.'}
                    )
            except ValueError:
                raise ValidationError({'date_to': 'Invalid date format. Use YYYY-MM-DD.'})

        return {
            'branch_id': branch_obj.id,
            'report_date': report_date,
            'status': request.query_params.get('status', 'all'),
            'parent_customer': request.query_params.get('parent_customer'),
            'child_customer': request.query_params.get('child_customer'),
            'date_mode': request.query_params.get('date_mode', 'transaction_date'),
            'date_from': request.query_params.get('date_from'),
            'date_to': date_to,
            'show_instrument_numbers': request.query_params.get(
                'show_instrument_numbers', 'false'
            ).lower() == 'true',
            'return_only': request.query_params.get('return_only', 'false').lower() == 'true',
        }

    def _execute_query(self, sql_template, sql_params, filter_snippets):
        sql = sql_template.format(**filter_snippets)
        with connection.cursor() as cursor:
            cursor.execute(sql, sql_params)
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        return rows

    def _get_data(self, params, use_query2=False):
        filter_snippets, extra_params = self._build_filters(params)
        query_template = self.QUERY2 if use_query2 else self.QUERY1
        sql_params = {'branch_id': params['branch_id'], 'report_date': params['report_date']}
        sql_params.update(extra_params)
        return self._execute_query(query_template, sql_params, filter_snippets)

    def _get_count_and_totals(self, params):
        filter_snippets, extra_params = self._build_filters(params)
        sql_params = {'branch_id': params['branch_id'], 'report_date': params.get('report_date', timezone.now().date().isoformat())}
        sql_params.update(extra_params)

        count_rows = self._execute_query(self.COUNT_QUERY, sql_params, filter_snippets)
        total_count = count_rows[0]['total'] if count_rows else 0

        total_rows = self._execute_query(self.TOTALS_QUERY, sql_params, filter_snippets)
        totals = total_rows[0] if total_rows else {'total_sales_amount': 0, 'total_sales_return': 0, 'total_net_sales': 0}

        return total_count, totals

    def get(self, request):
        try:
            params = self._get_common_params(request)
        except ValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)

        output = request.query_params.get('export', 'html').lower()
        if output == 'pdf':
            return self._export_pdf(request, params)
        elif output == 'excel':
            return self._export_excel(request, params)
        elif output == 'csv':
            return self._export_csv(request, params)
        return self._html_view(request, params)
    
    def fmt_date(self, value):
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime("%d-%b-%Y")
        try:
            dt = datetime.strptime(value, "%Y-%m-%d")
            return dt.strftime("%d-%b-%Y")
        except (TypeError, ValueError):
            return str(value)

    def _html_view(self, request, params):
        page_val = request.query_params.get('page', '1')
        try:
            page = int(page_val)
        except (ValueError, TypeError):
            page = 1
        page_size = int(request.query_params.get('page_size', 50))
        offset = (page - 1) * page_size

        use_query2 = params['show_instrument_numbers']
        filter_snippets, extra_params = self._build_filters(params)
        query_template = self.QUERY2 if use_query2 else self.QUERY1

        sql_params = {'branch_id': params['branch_id'], 'report_date': params['report_date']}
        sql_params.update(extra_params)
        sql_params['offset'] = offset
        sql_params['limit'] = page_size

        paginated_query = query_template + ' OFFSET %(offset)s LIMIT %(limit)s'
        rows = self._execute_query(paginated_query, sql_params, filter_snippets)
        total_count, totals = self._get_count_and_totals(params)
        total_pages = max(1, (total_count + page_size - 1) // page_size)

        return Response({
            'data': rows,
            'page': page,
            'page_size': page_size,
            'total_count': total_count,
            'total_pages': total_pages,
            'totals': totals,
            'filter_params': {
                'status': params['status'],
                'date_mode': params['date_mode'],
                'date_from': params.get('date_from'),
                'date_to': params.get('date_to'),
                'report_date': params['report_date'],
                'parent_customer': params.get('parent_customer'),
                'child_customer': params.get('child_customer'),
                'show_instrument_numbers': params['show_instrument_numbers'],
                'return_only': params['return_only'],
            }
        })

    def _export_pdf(self, request, params):
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm

        data = self._get_data(params, use_query2=False)
        total_count, totals = self._get_count_and_totals(params)

        groups = {}                   

        for row in data:
            pid = row['payment_sl']
            groups.setdefault(pid, []).append(row)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('Invoice and Payment Information', styles['Title']))
        elements.append(Spacer(1, 6*mm))

        filter_text = f"Report Date: {params['report_date']} | Status: {params['status']} | Date Mode: {params['date_mode']}"
        if params.get('date_from'):
            filter_text += f" | From: {params['date_from']}"
        if params.get('date_to'):
            filter_text += f" | To: {params['date_to']}"
        if params.get('parent_customer'):
            filter_text += f" | Parent: {params['parent_customer']}"
        if params.get('child_customer'):
            filter_text += f" | Customer: {params['child_customer']}"
        if params['return_only']:
            filter_text += ' | Return Only: Yes'
        elements.append(Paragraph(filter_text, styles['Normal']))
        elements.append(Spacer(1, 4*mm))

        header_cols = ['Received Date', 'Cheque/Cash Amt', 'Claim Amt', 'Shortage Amt']
        detail_cols = ['SL No', 'Parent Org', 'Customer', 'Trans Date', 'Grace Days', 'Due Date', 'Sales Amt', 'Sales Return', 'Net Sales', 'Days Overdue']

        grand_sales = grand_return = grand_net = 0

        for pid, rows in groups.items():
            first = rows[0]
            if pid is None:
                hdr_data = [['Unpaid', '', '', '']]
            else:
                hdr_data = [[self.fmt_date(first.get('received_date') or ''), str(first.get('cheque_cash_amount') or '0'), str(first.get('claim_amount') or '0'), str(first.get('shortage_amount') or '0')]]

            ht = Table([header_cols] + hdr_data)
            ht.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90d9')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e6f3ff')),
            ]))
            elements.append(ht)
            elements.append(Spacer(1, 2*mm))

            d_rows = [[r['sl_no'], r['parent_organization'] or '', r['customer_name'] or '', self.fmt_date(r['transaction_date']), str(r['grace_days']), self.fmt_date(r.get('due_date', r.get('Due_Date', '')) or ''), str(r['sales_amount']), str(r['sales_return']), str(r['net_sales']), str(r['days_overdue'])] for r in rows]
            dt = Table([detail_cols] + d_rows)
            dt.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
                ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ]))
            elements.append(dt)
            elements.append(Spacer(1, 2*mm))

            gs = sum(r['sales_amount'] for r in rows)
            gr = sum(r['sales_return'] for r in rows)
            gn = sum(r['net_sales'] for r in rows)
            grand_sales += gs
            grand_return += gr
            grand_net += gn

            ft = Table([['', '', '', '', 'Group Total:', str(gs), str(gr), str(gn), '']])
            ft.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d4edda')),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
                ('ALIGN', (4, 0), (6, 0), 'RIGHT'),
            ]))
            elements.append(ft)
            elements.append(Spacer(1, 4*mm))

        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph(f'Grand Totals - Sales Amount: {grand_sales}, Sales Return: {grand_return}, Net Sales: {grand_net}', styles['Normal']))
        doc.build(elements)
        buf.seek(0)

        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="credit_invoice_report.pdf"'
        return response

    def _export_excel(self, request, params):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        data = self._get_data(params, use_query2=True)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Credit Invoice Report'

        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        title_font = Font(name='Calibri', bold=True, size=14, color='2F5496')
        label_font = Font(name='Calibri', bold=True, size=10)
        value_font = Font(name='Calibri', size=10)

        thin_border = Border(
            left=Side(style='thin', color='B0B0B0'),
            right=Side(style='thin', color='B0B0B0'),
            top=Side(style='thin', color='B0B0B0'),
            bottom=Side(style='thin', color='B0B0B0'),
        )
        date_alignment = Alignment(horizontal='center', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        text_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        col_widths = {
            1: 8, 2: 28, 3: 28, 4: 16, 5: 12, 6: 16, 7: 16,
            8: 16, 9: 16, 10: 16, 11: 30, 12: 18, 13: 30, 14: 18,
            15: 16, 16: 14, 17: 14,
        }

        ws.merge_cells('A1:Q1')
        title_cell = ws['A1']
        title_cell.value = 'Credit Invoice Report'
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        parent_customer_alias = params.get('parent_customer')
        child_customer_alias = params.get('child_customer')
        parent_customer_name = 'All'
        child_customer_name = 'All'
        if parent_customer_alias:
            try:
                parent_customer_name = Customer.objects.values_list('name', flat=True).get(alias_id=parent_customer_alias)
            except Customer.DoesNotExist:
                pass
        if child_customer_alias:
            try:
                child_customer_name = Customer.objects.values_list('name', flat=True).get(alias_id=child_customer_alias)
            except Customer.DoesNotExist:
                pass

        date_from = params.get('date_from')
        date_to = params.get('date_to')
        date_from_str = self.fmt_date(date_from) if date_from else 'N/A'
        date_to_str = self.fmt_date(date_to) if date_to else 'N/A'

        filter_info = [
            ('Report Date', params['report_date']),
            ('Status', params['status']),
            ('Date Mode', params['date_mode']),
            ('Date From', date_from_str),
            ('Date To', date_to_str),
            ('Parent Customer', parent_customer_name),
            ('Child Customer', child_customer_name),
            ('Return Only', 'Yes' if params['return_only'] else 'No'),
        ]
        for i, (label, value) in enumerate(filter_info):
            row_num = 2 + i
            ws.cell(row=row_num, column=1, value=label).font = label_font
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=4)
            val_cell = ws.cell(row=row_num, column=2, value=value)
            val_cell.font = value_font
            val_cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.append([])

        show_instruments = params['show_instrument_numbers']
        if show_instruments:
            cols = ['SL No', 'Parent Organization', 'Customer Name', 'Transaction Date', 'Grace Days', 'Due Date', 'Received Date', 'Sales Amount', 'Sales Return', 'Net Sales', 'Cheque Numbers', 'Cheque/Cash Amount', 'Claim Numbers', 'Claim Amount', 'Shortage Amount', 'Days Overdue', 'Payment SL']
        else:
            cols = ['SL No', 'Parent Organization', 'Customer Name', 'Transaction Date', 'Grace Days', 'Due Date', 'Received Date', 'Sales Amount', 'Sales Return', 'Net Sales', 'Cheque/Cash Amount', 'Claim Amount', 'Shortage Amount', 'Days Overdue', 'Payment SL']

        header_row = ws.max_row + 1
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 28

        for row in data:
            trans_date = self.fmt_date(row.get("transaction_date"))
            recv_date = self.fmt_date(row.get("received_date"))
            due_date = self.fmt_date(row.get("Due_Date") or row.get("due_date"))
            data_row_num = ws.max_row + 1

            common_values = [
                (1, int(row.get("sl_no", 0)), number_alignment),
                (2, row.get("parent_organization", ""), text_alignment),
                (3, row.get("customer_name", ""), text_alignment),
                (4, trans_date, date_alignment),
                (5, int(row.get("grace_days", 0)), number_alignment),
                (6, due_date, date_alignment),
                (7, recv_date, date_alignment),
            ]

            if show_instruments:
                values = common_values + [
                    (8, float(row.get("sales_amount", 0)), number_alignment),
                    (9, float(row.get("sales_return", 0)), number_alignment),
                    (10, float(row.get("net_sales", 0)), number_alignment),
                    (11, row.get("cheque_numbers", ""), text_alignment),
                    (12, float(row.get("cheque_cash_amount", 0) or 0), number_alignment),
                    (13, row.get("claim_numbers", ""), text_alignment),
                    (14, float(row.get("claim_amount", 0) or 0), number_alignment),
                    (15, float(row.get("shortage_amount", 0) or 0), number_alignment),
                    (16, int(row.get("days_overdue", 0)), number_alignment),
                    (17, int(row.get('payment_sl', 0)), number_alignment),
                ]
            else:
                values = common_values + [
                    (8, float(row.get("sales_amount", 0)), number_alignment),
                    (9, float(row.get("sales_return", 0)), number_alignment),
                    (10, float(row.get("net_sales", 0)), number_alignment),
                    (11, float(row.get("cheque_cash_amount", 0) or 0), number_alignment),
                    (12, float(row.get("claim_amount", 0) or 0), number_alignment),
                    (13, float(row.get("shortage_amount", 0) or 0), number_alignment),
                    (14, int(row.get("days_overdue", 0)), number_alignment),
                    (15, int(row.get('payment_sl', 0)), number_alignment),
                ]

            for col_idx, val, align in values:
                cell = ws.cell(row=data_row_num, column=col_idx, value=val)
                cell.font = value_font
                cell.alignment = align
                cell.border = thin_border

        total_cols = len(cols)
        for col_idx in range(1, total_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 14)

        ws.freeze_panes = f'A{header_row + 1}'

        last_col_letter = get_column_letter(total_cols)
        ws.auto_filter.ref = f'A{header_row}:{last_col_letter}{ws.max_row}'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        wb.save(response)
        response['Content-Disposition'] = 'attachment; filename="credit_invoice_report.xlsx"'
        return response

    def _export_csv(self, request, params):
        import csv
        data = self._get_data(params, use_query2=True)

        def csv_generator():
            yield '# Credit Invoice Report\n'
            yield f'# Report Date: {params["report_date"]}\n'
            yield f'# Status: {params["status"]}\n'
            yield f'# Date Mode: {params["date_mode"]}\n'
            yield f'# Date From: {params.get("date_from", "N/A")}\n'
            yield f'# Date To: {params.get("date_to", "N/A")}\n'
            yield f'# Parent Customer: {params.get("parent_customer", "All")}\n'
            yield f'# Child Customer: {params.get("child_customer", "All")}\n'
            yield f'# Return Only: {"Yes" if params["return_only"] else "No"}\n'
            yield '\n'

            if params['show_instrument_numbers']:
                cols = ['Invoice No', 'Parent Organization', 'Customer Name', 'Transaction Date', 'Grace Days', 'Due Date', 'Received Date', 'Sales Amount', 'Sales Return', 'Net Sales', 'Cheque Numbers', 'Cheque/Cash Amount', 'Claim Numbers', 'Claim Amount', 'Shortage Amount', 'Payment SL', 'Days Overdue']
            else:
                cols = ['Invoice No', 'Parent Organization', 'Customer Name', 'Transaction Date', 'Grace Days', 'Due Date', 'Received Date', 'Sales Amount', 'Sales Return', 'Net Sales', 'Cheque/Cash Amount', 'Claim Amount', 'Shortage Amount', 'Payment SL', 'Days Overdue']

            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(cols)
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)

            for row in data:
                if params['show_instrument_numbers']:
                    writer.writerow([row.get("sl_no") or "", row.get("parent_organization") or "", row.get("customer_name") or "", self.fmt_date(row.get("transaction_date")), row.get("grace_days") or 0, self.fmt_date(row.get("Due_Date") or row.get("due_date")), self.fmt_date(row.get("received_date")), row.get("sales_amount") or 0, row.get("sales_return") or 0, row.get("net_sales") or 0, row.get("cheque_numbers") or "", row.get("cheque_cash_amount") or 0, row.get("claim_numbers") or "", row.get("claim_amount") or 0, row.get("shortage_amount") or 0, str(row.get('payment_sl') or ''), row.get("days_overdue") or 0])
                else:
                    writer.writerow([row.get("sl_no") or "", row.get("parent_organization") or "", row.get("customer_name") or "", self.fmt_date(row.get("transaction_date")), row.get("grace_days") or 0, self.fmt_date(row.get("Due_Date") or row.get("due_date")), self.fmt_date(row.get("received_date")), row.get("sales_amount") or 0, row.get("sales_return") or 0, row.get("net_sales") or 0, row.get("cheque_cash_amount") or 0, row.get("claim_amount") or 0, row.get("shortage_amount") or 0, str(row.get('payment_sl') or ''), row.get("days_overdue") or 0])
                yield buf.getvalue()
                buf.seek(0)
                buf.truncate(0)

        response = StreamingHttpResponse(csv_generator(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="credit_invoice_report.csv"'
        return response


# ==================== Received and Claim Report Module ====================

class ReceivedClaimReportView(APIView):
    '''
    Received and Claim Report with filtering, pagination, PDF, Excel exports.
    Endpoint: GET /v1/chq/reports/received-claim/
    '''
    permission_classes = [IsAuthenticated]

    RECEIVE_QUERY = '''
        SELECT
            ROW_NUMBER() OVER (ORDER BY p.received_date, pd.id) AS sl_no,
            p.alias_id AS payment_id,
            p.received_date,
            c.name AS organization_name,
            pi.instrument_name,
            pd.id_number AS instrument_number,
            pd.amount,
            pd.detail AS remarks
        FROM payment_details pd
        JOIN payment p ON p.id = pd.payment_id
        JOIN customer c ON c.id = p.customer_id
        LEFT JOIN customer pc ON pc.id = c.parent_id AND pc.is_parent = TRUE
        JOIN payment_instrument pi ON pi.id = pd.payment_instrument_id
        WHERE p.branch_id = %(branch_id)s
        {parent_filter}
        {instrument_filter}
        {date_filter}
        {comments_filter}
        ORDER BY p.received_date, pd.id
    '''

    CLAIM_QUERY = '''
        SELECT
            ROW_NUMBER() OVER (ORDER BY p.received_date, cl.id) AS sl_no,
            cl.alias_id AS claim_id,
            p.received_date AS claim_date,
            c.name AS organization_name,
            pi.instrument_name,
            pd.id_number AS instrument_number,
            pd.amount,
            cl.refund_amount,
            cl.refund_date,
            (pd.amount - COALESCE(cl.refund_amount, 0)) AS remaining_amount,
            cl.remarks
        FROM claim cl
        JOIN payment_details pd ON pd.id = cl.payment_details_id
        JOIN payment p ON p.id = pd.payment_id
        JOIN customer c ON c.id = p.customer_id
        LEFT JOIN customer pc ON pc.id = c.parent_id AND pc.is_parent = TRUE
        JOIN payment_instrument pi ON pi.id = pd.payment_instrument_id
        WHERE p.branch_id = %(branch_id)s
        {parent_filter}
        {instrument_filter}
        {date_filter}
        {refund_filter}
        {comments_filter}
        ORDER BY p.received_date, cl.id
    '''

    RECEIVE_COUNT_QUERY = '''
        SELECT COUNT(*) AS total
        FROM payment_details pd
        JOIN payment p ON p.id = pd.payment_id
        JOIN customer c ON c.id = p.customer_id
        LEFT JOIN customer pc ON pc.id = c.parent_id AND pc.is_parent = TRUE
        JOIN payment_instrument pi ON pi.id = pd.payment_instrument_id
        WHERE p.branch_id = %(branch_id)s
        {parent_filter}
        {instrument_filter}
        {date_filter}
        {comments_filter}
    '''

    CLAIM_COUNT_QUERY = '''
        SELECT COUNT(*) AS total
        FROM claim cl
        JOIN payment_details pd ON pd.id = cl.payment_details_id
        JOIN payment p ON p.id = pd.payment_id
        JOIN customer c ON c.id = p.customer_id
        LEFT JOIN customer pc ON pc.id = c.parent_id AND pc.is_parent = TRUE
        JOIN payment_instrument pi ON pi.id = pd.payment_instrument_id
        WHERE p.branch_id = %(branch_id)s
        {parent_filter}
        {instrument_filter}
        {date_filter}
        {refund_filter}
        {comments_filter}
    '''

    RECEIVE_TOTALS_QUERY = '''
        SELECT
            COALESCE(SUM(pd.amount), 0) AS total_amount
        FROM payment_details pd
        JOIN payment p ON p.id = pd.payment_id
        JOIN customer c ON c.id = p.customer_id
        LEFT JOIN customer pc ON pc.id = c.parent_id AND pc.is_parent = TRUE
        JOIN payment_instrument pi ON pi.id = pd.payment_instrument_id
        WHERE p.branch_id = %(branch_id)s
        {parent_filter}
        {instrument_filter}
        {date_filter}
        {comments_filter}
    '''

    CLAIM_TOTALS_QUERY = '''
        SELECT
            COALESCE(SUM(pd.amount), 0) AS total_amount,
            COALESCE(SUM(cl.refund_amount), 0) AS total_refund,
            COALESCE(SUM(pd.amount - COALESCE(cl.refund_amount, 0)), 0) AS total_remaining
        FROM claim cl
        JOIN payment_details pd ON pd.id = cl.payment_details_id
        JOIN payment p ON p.id = pd.payment_id
        JOIN customer c ON c.id = p.customer_id
        LEFT JOIN customer pc ON pc.id = c.parent_id AND pc.is_parent = TRUE
        JOIN payment_instrument pi ON pi.id = pd.payment_instrument_id
        WHERE p.branch_id = %(branch_id)s
        {parent_filter}
        {instrument_filter}
        {date_filter}
        {refund_filter}
        {comments_filter}
    '''

    def _build_filters(self, params, report_type):
        snippets = {
            'parent_filter': '',
            'instrument_filter': '',
            'date_filter': '',
            'refund_filter': '',
            'comments_filter': '',
        }
        extra_params = {}

        parent_customer = params.get('parent_customer')
        if parent_customer:
            snippets['parent_filter'] = " AND c.alias_id = %(parent_customer)s"
            extra_params['parent_customer'] = parent_customer

        instrument_ids = params.get('instrument_ids')
        if instrument_ids:
            id_list = [x.strip() for x in instrument_ids.split(',') if x.strip()]
            if id_list:
                placeholders = ','.join([f'%(inst_{i})s' for i in range(len(id_list))])
                snippets['instrument_filter'] = f" AND pi.id IN ({placeholders})"
                for i, val in enumerate(id_list):
                    extra_params[f'inst_{i}'] = val

        if report_type == 'receive':
            date_from = params.get('date_from')
            date_to = params.get('date_to')
            if date_from and date_to:
                snippets['date_filter'] = " AND p.received_date BETWEEN %(date_from)s::date AND %(date_to)s::date"
                extra_params['date_from'] = date_from
                extra_params['date_to'] = date_to
            elif date_from:
                snippets['date_filter'] = " AND p.received_date >= %(date_from)s::date"
                extra_params['date_from'] = date_from
            elif date_to:
                snippets['date_filter'] = " AND p.received_date <= %(date_to)s::date"
                extra_params['date_to'] = date_to

            if params.get('has_comments') in (True, 'true', 'True', 1, '1'):
                snippets['comments_filter'] = " AND pd.detail IS NOT NULL AND pd.detail != ''"

        elif report_type == 'claim':
            date_mode = params.get('date_mode', 'received_date')
            date_from = params.get('date_from')
            date_to = params.get('date_to')

            date_col = 'cl.refund_date' if date_mode == 'refund_date' else 'p.received_date'
            if date_from and date_to:
                snippets['date_filter'] = f" AND {date_col} BETWEEN %(date_from)s::date AND %(date_to)s::date"
                extra_params['date_from'] = date_from
                extra_params['date_to'] = date_to
            elif date_from:
                snippets['date_filter'] = f" AND {date_col} >= %(date_from)s::date"
                extra_params['date_from'] = date_from
            elif date_to:
                snippets['date_filter'] = f" AND {date_col} <= %(date_to)s::date"
                extra_params['date_to'] = date_to

            refund_status = params.get('refund_status', 'all')
            if refund_status == 'refunded':
                snippets['refund_filter'] = " AND cl.refund_amount > 0"
            elif refund_status == 'pending':
                snippets['refund_filter'] = " AND (cl.refund_amount IS NULL OR cl.refund_amount = 0)"

            if params.get('has_comments') in (True, 'true', 'True', 1, '1'):
                snippets['comments_filter'] = " AND cl.remarks IS NOT NULL AND cl.remarks != ''"

        return snippets, extra_params

    def _get_common_params(self, request):
        branch = request.query_params.get('branch')
        if not branch:
            raise ValidationError({'branch': 'Branch is required.'})

        branch_obj = get_object_or_404(Branch, alias_id=branch)

        return {
            'branch_id': branch_obj.id,
            'report_type': request.query_params.get('report_type', 'receive'),
            'parent_customer': request.query_params.get('parent_customer'),
            'instrument_ids': request.query_params.get('instrument_ids'),
            'date_from': request.query_params.get('date_from'),
            'date_to': request.query_params.get('date_to'),
            'date_mode': request.query_params.get('date_mode', 'received_date'),
            'refund_status': request.query_params.get('refund_status', 'all'),
            'has_comments': request.query_params.get('has_comments', 'false').lower() == 'true',
        }

    def _execute_query(self, sql_template, sql_params, filter_snippets):
        sql = sql_template.format(**filter_snippets)
        with connection.cursor() as cursor:
            cursor.execute(sql, sql_params)
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        return rows

    def _get_data(self, params):
        filter_snippets, extra_params = self._build_filters(params, params['report_type'])
        query_template = self.RECEIVE_QUERY if params['report_type'] == 'receive' else self.CLAIM_QUERY
        sql_params = {'branch_id': params['branch_id']}
        sql_params.update(extra_params)
        return self._execute_query(query_template, sql_params, filter_snippets)

    def _get_count_and_totals(self, params):
        filter_snippets, extra_params = self._build_filters(params, params['report_type'])
        is_receive = params['report_type'] == 'receive'
        count_query = self.RECEIVE_COUNT_QUERY if is_receive else self.CLAIM_COUNT_QUERY
        totals_query = self.RECEIVE_TOTALS_QUERY if is_receive else self.CLAIM_TOTALS_QUERY

        sql_params = {'branch_id': params['branch_id']}
        sql_params.update(extra_params)

        count_rows = self._execute_query(count_query, sql_params, filter_snippets)
        total_count = count_rows[0]['total'] if count_rows else 0

        total_rows = self._execute_query(totals_query, sql_params, filter_snippets)
        totals = total_rows[0] if total_rows else {}

        return total_count, totals

    def get(self, request):
        try:
            params = self._get_common_params(request)
        except ValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)

        output = request.query_params.get('export', 'html').lower()
        if output == 'pdf':
            return self._export_pdf(request, params)
        elif output == 'excel':
            return self._export_excel(request, params)
        return self._html_view(request, params)

    def fmt_date(self, value):
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime("%d-%b-%Y")
        try:
            dt = datetime.strptime(value, "%Y-%m-%d")
            return dt.strftime("%d-%b-%Y")
        except (TypeError, ValueError):
            return str(value)

    def _html_view(self, request, params):
        page_val = request.query_params.get('page', '1')
        try:
            page = int(page_val)
        except (ValueError, TypeError):
            page = 1
        page_size = int(request.query_params.get('page_size', 50))
        offset = (page - 1) * page_size

        filter_snippets, extra_params = self._build_filters(params, params['report_type'])
        query_template = self.RECEIVE_QUERY if params['report_type'] == 'receive' else self.CLAIM_QUERY

        sql_params = {'branch_id': params['branch_id']}
        sql_params.update(extra_params)
        sql_params['offset'] = offset
        sql_params['limit'] = page_size

        paginated_query = query_template + ' OFFSET %(offset)s LIMIT %(limit)s'
        rows = self._execute_query(paginated_query, sql_params, filter_snippets)
        total_count, totals = self._get_count_and_totals(params)
        total_pages = max(1, (total_count + page_size - 1) // page_size)

        return Response({
            'data': rows,
            'page': page,
            'page_size': page_size,
            'total_count': total_count,
            'total_pages': total_pages,
            'totals': totals,
            'filter_params': {
                'report_type': params['report_type'],
                'parent_customer': params.get('parent_customer'),
                'instrument_ids': params.get('instrument_ids'),
                'date_from': params.get('date_from'),
                'date_to': params.get('date_to'),
                'date_mode': params.get('date_mode'),
                'refund_status': params.get('refund_status'),
                'has_comments': params.get('has_comments'),
            }
        })

    def _export_pdf(self, request, params):
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm

        data = self._get_data(params)
        total_count, totals = self._get_count_and_totals(params)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        elements = []

        title = 'Received Report' if params['report_type'] == 'receive' else 'Claim Report'
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 6*mm))

        filter_text = f"Report Type: {params['report_type']}"
        if params.get('parent_customer'):
            filter_text += f" | Parent: {params['parent_customer']}"
        if params.get('date_from'):
            filter_text += f" | From: {params['date_from']}"
        if params.get('date_to'):
            filter_text += f" | To: {params['date_to']}"
        elements.append(Paragraph(filter_text, styles['Normal']))
        elements.append(Spacer(1, 4*mm))

        if params['report_type'] == 'receive':
            cols = ['SL No', 'Received Date', 'Organization', 'Instrument', 'Instrument No', 'Amount', 'Remarks']
            table_data = [[r['sl_no'], self.fmt_date(r['received_date']), r['organization_name'] or '', r['instrument_name'] or '', r['instrument_number'] or '', str(r['amount']), r['remarks'] or ''] for r in data]
        else:
            cols = ['SL No', 'Claim Date', 'Organization', 'Instrument', 'Instrument No', 'Amount', 'Refund', 'Refund Date', 'Refund Amt', 'Remaining', 'Remarks']
            table_data = [[r['sl_no'], self.fmt_date(r['claim_date']), r['organization_name'] or '', r['instrument_name'] or '', r['instrument_number'] or '', str(r['amount']), 'Yes' if r['refund_amount'] and float(r['refund_amount']) > 0 else 'No', self.fmt_date(r['refund_date']), str(r['refund_amount'] or 0), str(r['remaining_amount'] or 0), r['remarks'] or ''] for r in data]

        header = [cols]
        body = header + table_data
        # Footer row
        if params['report_type'] == 'receive' and totals:
            footer = ['', '', '', '', 'Total:', str(totals.get('total_amount', 0)), '']
            body.append(footer)
        elif totals:
            footer = ['', '', '', '', 'Total:', str(totals.get('total_amount', 0)), '', '', str(totals.get('total_refund', 0)), str(totals.get('total_remaining', 0)), '']
            body.append(footer)

        t = Table(body)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d4edda')),
        ]))
        elements.append(t)
        doc.build(elements)
        buf.seek(0)

        filename = 'received_report.pdf' if params['report_type'] == 'receive' else 'claim_report.pdf'
        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _export_excel(self, request, params):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        data = self._get_data(params)
        total_count, totals = self._get_count_and_totals(params)

        wb = Workbook()
        ws = wb.active
        title = 'Received Report' if params['report_type'] == 'receive' else 'Claim Report'
        ws.title = title

        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        title_font = Font(name='Calibri', bold=True, size=14, color='2F5496')
        label_font = Font(name='Calibri', bold=True, size=10)
        value_font = Font(name='Calibri', size=10)

        thin_border = Border(
            left=Side(style='thin', color='B0B0B0'),
            right=Side(style='thin', color='B0B0B0'),
            top=Side(style='thin', color='B0B0B0'),
            bottom=Side(style='thin', color='B0B0B0'),
        )
        date_alignment = Alignment(horizontal='center', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        text_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        max_col = 7 if params['report_type'] == 'receive' else 11
        last_col_letter = get_column_letter(max_col)

        ws.merge_cells(f'A1:{last_col_letter}1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        parent_customer_alias = params.get('parent_customer')
        parent_customer_name = 'All'
        if parent_customer_alias:
            try:
                parent_customer_name = Customer.objects.values_list('name', flat=True).get(alias_id=parent_customer_alias)
            except Customer.DoesNotExist:
                pass

        filter_info = [
            ('Report Type', params['report_type']),
            ('Parent Customer', parent_customer_name),
            ('Date From', params.get('date_from', 'N/A')),
            ('Date To', params.get('date_to', 'N/A')),
        ]
        if params['report_type'] == 'claim':
            filter_info.insert(1, ('Date Mode', params.get('date_mode', 'received_date')))
            filter_info.insert(2, ('Refund Status', params.get('refund_status', 'all')))

        for i, (label, value) in enumerate(filter_info):
            row_num = 2 + i
            ws.cell(row=row_num, column=1, value=label).font = label_font
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=4)
            val_cell = ws.cell(row=row_num, column=2, value=value)
            val_cell.font = value_font
            val_cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.append([])

        if params['report_type'] == 'receive':
            cols = ['SL No', 'Received Date', 'Organization Name', 'Instrument Name', 'Instrument Number', 'Amount', 'Remarks']
            col_widths = {1: 8, 2: 16, 3: 28, 4: 22, 5: 18, 6: 16, 7: 30}
        else:
            cols = ['SL No', 'Claim Date', 'Organization Name', 'Instrument Name', 'Instrument Number', 'Amount', 'Refund', 'Refund Date', 'Refund Amount', 'Remaining Amount', 'Remarks']
            col_widths = {1: 8, 2: 16, 3: 28, 4: 22, 5: 18, 6: 16, 7: 10, 8: 16, 9: 16, 10: 18, 11: 30}

        header_row = ws.max_row + 1
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 28

        for row in data:
            data_row_num = ws.max_row + 1
            if params['report_type'] == 'receive':
                values = [
                    (1, int(row.get('sl_no', 0)), number_alignment),
                    (2, self.fmt_date(row.get('received_date')), date_alignment),
                    (3, row.get('organization_name', ''), text_alignment),
                    (4, row.get('instrument_name', ''), text_alignment),
                    (5, row.get('instrument_number', ''), text_alignment),
                    (6, float(row.get('amount', 0)), number_alignment),
                    (7, row.get('remarks', ''), text_alignment),
                ]
            else:
                values = [
                    (1, int(row.get('sl_no', 0)), number_alignment),
                    (2, self.fmt_date(row.get('claim_date')), date_alignment),
                    (3, row.get('organization_name', ''), text_alignment),
                    (4, row.get('instrument_name', ''), text_alignment),
                    (5, row.get('instrument_number', ''), text_alignment),
                    (6, float(row.get('amount', 0)), number_alignment),
                    (7, 'Yes' if row.get('refund_amount') and float(row['refund_amount']) > 0 else 'No', text_alignment),
                    (8, self.fmt_date(row.get('refund_date')), date_alignment),
                    (9, float(row.get('refund_amount', 0) or 0), number_alignment),
                    (10, float(row.get('remaining_amount', 0) or 0), number_alignment),
                    (11, row.get('remarks', ''), text_alignment),
                ]

            for col_idx, val, align in values:
                cell = ws.cell(row=data_row_num, column=col_idx, value=val)
                cell.font = value_font
                cell.alignment = align
                cell.border = thin_border

        for col_idx in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 14)

        ws.freeze_panes = f'A{header_row + 1}'
        ws.auto_filter.ref = f'A{header_row}:{last_col_letter}{ws.max_row}'

        filename = 'received_report.xlsx' if params['report_type'] == 'receive' else 'claim_report.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        wb.save(response)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response